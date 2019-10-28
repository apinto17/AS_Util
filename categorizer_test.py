import pandas as pd
import re
import numpy as np
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.svm import LinearSVC
from sklearn.svm import SVC
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import SelectKBest, chi2
from sklearn.feature_extraction.text import HashingVectorizer
from sklearn.preprocessing import MaxAbsScaler
from sklearn.calibration import CalibratedClassifierCV
import pickle
from openpyxl import load_workbook
import os.path
import xlsxwriter
from xlrd import open_workbook
import xlwt
import sys

#Added by Ian-9/17:
import mysql.connector
import sshtunnel

NUM_SUG = 5
NUM_CATS = 941
NUM_AUTO_SAVE = 10
selections = []
cat_strings = []


# to test use Abrasives|Abrasive Brushes|Abrasive Brush Accessories


def pickle_model():

    stemmer = SnowballStemmer("english")
    words = stopwords.words("english")

    data = pd.read_excel("grainger_cleaned_w_o_dups.xlsx")

    #create train/test split and adjust settings
    X_train, X_test, Y_train, Y_test = train_test_split(data["cleaned"], data.output_category, test_size=.10)

    svm = LinearSVC(random_state=0, max_iter=10000, C=.4)
    clf = CalibratedClassifierCV(svm)
    vect = CountVectorizer(ngram_range=(1,2))
    pipeline = Pipeline([
            ("vect", vect),
            ("clf", clf)
    ])

    print("training...")
    model = pipeline.fit(data["cleaned"], data["output_category"])

    #write model to pickle
    model_file = open("model_pickle", "wb")
    pickle.dump(model, model_file)

    model_file.close()


def Select_Input_Cat_From_DB():
    #ssh tunnel
    sshtunnel.SSH_TIMEOUT = 350.0
    sshtunnel.TUNNEL_TIMEOUT = 350.0

    with sshtunnel.SSHTunnelForwarder('ssh.pythonanywhere.com',
                                      ssh_username='iclam19',
                                      ssh_password='@astest@1234',
                                      remote_bind_address=('iclam19.mysql.pythonanywhere-services.com'
                                      , 3306)) as tunnel:
        connection = mysql.connector.connect(user='iclam19',
                password='astest1234', host='127.0.0.1',
                port=tunnel.local_bind_port,
                database='iclam19$AssembledSupply')
        mycursor = connection.cursor()
        sql = "Select input_category From ft_ml_categories Where output_category is null  Order by id asc limit 100"

        mycursor.execute(sql)
        returned_items = list(mycursor.fetchall())
        connection.close()
        return returned_items



def run_categorizer():

    res = []
    data = pd.read_excel("grainger_cleaned_w_o_dups.xlsx")
    model_file = open("model_pickle", "rb")
    model = pickle.load(model_file)
    model_file.close()
    #new_session_I_C = []
    new_session_I_C = Select_Input_Cat_From_DB()
    i_c = [i_c[0] for i_c in new_session_I_C]
    for x in i_c:
    #cat_string = input()
        cat_string = x

        while(True):
            if(len(selections) > NUM_AUTO_SAVE):
                save()
            print("---------------------------------------------------------")
            print("Enter category string or type \'exit\' to save and exit ")
            print("type \'s\' to skip and \'f\' to add five more")
            print("!!!IMPORTANT!!! TYPE EXIT DONT JUST EXIT")
            print('Categorize the following:', cat_string)

            if(cat_string == "exit"):
                save_and_exit()
            get_suggestions(cat_string, data, model)


def get_suggestions(cat_string, data, model):

    best_n_indx = 0
    res_num = 1
    res = []

    # grab initial data
    res.append((res_num, str(model.predict([cat_string])), "Unknown"))
    res_num += 1
    probs = model.predict_proba([cat_string])
    probs = probs.flatten()
    best_n = np.argsort(probs)[::-1]

    res, best_n_indx, res_num, selection = display_data(res, best_n, probs, best_n_indx, res_num, data)

    # if user printed 5 more
    while(type(selection) is str and selection == "f"):
        res, best_n_indx, res_num, selection = display_data(res, best_n, probs, best_n_indx, res_num, data)

    append_selection(cat_string, selection, res)


def display_data(res, best_n, probs, best_n_indx, res_num, data):

    limit = NUM_SUG - 1 if best_n_indx == 0 else NUM_SUG
    for i in range(limit):
        res.append((res_num, data["output_unique"].iloc[best_n[best_n_indx]],probs[best_n[best_n_indx]]))
        best_n_indx += 1
        res_num += 1

    print_res(res, res_num)
    selection = handle_input()

    return (res, best_n_indx, res_num, selection)


def append_selection(cat_string, selection, res):

    global selections
    global cat_strings

    if(type(selection) is int and selection > 0 and selection <= NUM_CATS):
        print("You Selected \"" + res[selection - 1][1] + "\"")
        selections.append(res[selection - 1][1])
        cat_strings.append(cat_string)


def handle_input():
    selection = input()
    selection = selection.lower()
    if(selection == "exit"):
        save_and_exit()
    elif(selection.isdigit()):
        selection = int(selection)
    return selection



def print_res(res, res_num):
    for i in range(-NUM_SUG + res_num - 1, res_num - 1):
        if(type(res[i][2]) is not str):
            print("          (" + str(res[i][0]) + ") " + str(res[i][1]) + "     Accuracy: %.2f%%" % (float(res[i][2]) * 100))
        else:
            print("          (" + str(res[i][0]) + ") " + str(res[i][1]) + "     Accuracy: " + res[i][2])


def save_and_exit():
    print("#########  EXITING: Writing results to \'selections.xlsx\'  #################")
    print(cat_strings)
    print(selections)

    df = pd.DataFrame({'input' : cat_strings,
                        'output' : selections})

    append_df_to_excel("selections.xlsx", df)

    exit()

def save():

    print(cat_strings)
    print(selections)

    df = pd.DataFrame({'input' : cat_strings,
                        'output' : selections})

    append_df_to_excel("selections.xlsx", df)
    print("#########  AUTO SAVE: Writing results to \'selections.xlsx\'  #################")



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        book = open_workbook(filename)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_wb():

    book = None
    fname = "selections.xlsx"
    if os.path.isfile(fname):
        book=load_workbook(fname)
    else:
        workbook2=xlwt.Workbook(fname)
        ws = workbook2.add_sheet('Sheet1')
        workbook2.save(fname)

    return book

if(__name__ == "__main__"):
    if(len(sys.argv) < 3):
        globals()[sys.argv[1]]()
    else:
        globals()[sys.argv[1]](sys.argv[2])




# def Select_Input_Cat_From_DB():
#     #ssh tunnel
#     sshtunnel.SSH_TIMEOUT = 350.0
#     sshtunnel.TUNNEL_TIMEOUT = 350.0

#     with sshtunnel.SSHTunnelForwarder('ssh.pythonanywhere.com',
#                                       ssh_username='iclam19',
#                                       ssh_password='@astest@1234',
#                                       remote_bind_address=('iclam19.mysql.pythonanywhere-services.com'
#                                       , 3306)) as tunnel:
#         connection = mysql.connector.connect(user='iclam19',
#                 password='astest1234', host='127.0.0.1',
#                 port=tunnel.local_bind_port,
#                 database='iclam19$AssembledSupply')
#         mycursor = connection.cursor()
#         sql = "Select input_category From ft_ml_categories Where output_category is null  Order by id asc limit 100"

#         mycursor.execute(sql)
#         returned_items = list(mycursor.fetchall())
#         return returned_items



def Update_PA_DB_w_Session_Data():
    df = pd.read_excel('selections.xlsx', sheet_name='Sheet1')

    for index, row in df.iterrows():
        #print(row['input'], row['output'])
        i_c = row['input']
        o_c = row['output']
        val = (str(o_c), str(i_c)) #oc is first for Set and ic is 2nd for Where

        #Insert Starts Here:
        sshtunnel.SSH_TIMEOUT = 350.0
        sshtunnel.TUNNEL_TIMEOUT = 350.0

        with sshtunnel.SSHTunnelForwarder(
            ('ssh.pythonanywhere.com'),
            ssh_username='iclam19', ssh_password='@astest@1234',
            remote_bind_address=('iclam19.mysql.pythonanywhere-services.com', 3306)
        ) as tunnel:
            connection = mysql.connector.connect(
                user='iclam19', password='astest1234',
                host='127.0.0.1', port=tunnel.local_bind_port,
                database='iclam19$AssembledSupply',
            )
            mycursor = connection.cursor()
            sql = "Update ft_ml_categories Set output_category = %s Where input_category = %s"
            mycursor.execute(sql, val)
            connection.commit()
            print(mycursor.rowcount, "# of record(s) updated")
            connection.close()

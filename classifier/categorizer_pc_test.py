import pandas as pd
import re
import numpy as np
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.svm import LinearSVC
from sklearn.svm import SVC
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import SelectKBest, chi2
from sklearn.feature_extraction.text import HashingVectorizer
from sklearn.preprocessing import MaxAbsScaler
from sklearn.calibration import CalibratedClassifierCV
import pickle
from openpyxl import load_workbook
import os.path
import xlsxwriter
from xlrd import open_workbook
import xlwt
import sys
from mysql.connector import Error
from mysql.connector import errorcode

import mysql.connector
import sshtunnel

NUM_SUG = 5
NUM_CATS = 941
NUM_AUTO_SAVE = 10
selections = []
cat_strings = []
server = None

# to test use Abrasives|Abrasive Brushes|Abrasive Brush Accessories


def train_model():

    global server

    pc = input("Enter a Primary_Category: ")
    server = Server(pc)

    stemmer = SnowballStemmer("english")
    words = stopwords.words("english")

    output_unique = server.Select_Unique_Output_Cat_From_DB()
    input_categories, output_categories = server.Select_Training_Data_From_DB()
    left_over = len(output_categories) - len(output_unique)
    for i in range(left_over):
        output_unique.append(None)


    data = pd.DataFrame({'input_category' : input_categories,
                        'output_category' : output_categories,
                        'output_unique' : output_unique})
    data["cleaned"] = data["input_category"].apply(lambda x: " ".join([stemmer.stem(i) for i in re.sub("[^a-zA-Z]", " ", x).split() if i not in words]).lower())

    svm = LinearSVC(random_state=0, max_iter=10000, C=.4)
    clf = CalibratedClassifierCV(svm)
    vect = CountVectorizer(ngram_range=(1,2))
    pipeline = Pipeline([
            ("vect", vect),
            ("clf", clf)
    ])

    print("training...")
    model = pipeline.fit(data["cleaned"], data["output_category"])

    return (model, data)


def run_categorizer():

    model, data = train_model()
    res = []
    new_session_I_C = server.Select_Input_Cat_From_DB()
    for x in new_session_I_C:
        cat_string = x[0]
        if(len(selections) > NUM_AUTO_SAVE):
            save()
        print("---------------------------------------------------------")
        print("Enter category string or type \'exit\' to save and exit ")
        print("type \'s\' to skip and \'f\' to add five more")
        print('Categorize the following:', cat_string)

        get_suggestions(cat_string, data, model)


def get_suggestions(cat_string, data, model):

    best_n_indx = 0
    res_num = 1
    res = []

    # grab initial data
    res.append((res_num, str(model.predict([cat_string])), "Unknown"))
    res_num += 1
    probs = model.predict_proba([cat_string])
    probs = probs.flatten()
    best_n = np.argsort(probs)[::-1]

    res, best_n_indx, res_num, selection = display_data(res, best_n, probs, best_n_indx, res_num, data)

    # if user printed 5 more
    while(type(selection) is str and selection == "f"):
        res, best_n_indx, res_num, selection = display_data(res, best_n, probs, best_n_indx, res_num, data)

    append_selection(cat_string, selection, res)


def display_data(res, best_n, probs, best_n_indx, res_num, data):

    limit = 0
    if best_n_indx == 0:
        limit = NUM_SUG - 1
    elif(len(best_n) <= (best_n_indx + NUM_SUG)):
        limit = len(best_n) - (best_n_indx + NUM_SUG)
    else:
        limit = NUM_SUG

    for i in range(limit):
        res.append((res_num, data["output_unique"].iloc[best_n[best_n_indx]],probs[best_n[best_n_indx]]))
        best_n_indx += 1
        res_num += 1

    print_res(res, res_num)
    selection = handle_input()

    return (res, best_n_indx, res_num, selection)


def append_selection(cat_string, selection, res):

    global selections
    global cat_strings

    if(type(selection) is int and selection > 0 and selection <= NUM_CATS):
        print("You Selected \"" + res[selection - 1][1] + "\"")
        selections.append(res[selection - 1][1])
        cat_strings.append(cat_string)


def handle_input():
    selection = input()
    selection = selection.lower()
    if(selection == "exit"):
        save_and_exit()
    elif(selection.isdigit()):
        selection = int(selection)
    return selection



def print_res(res, res_num):
    for i in range(-NUM_SUG + res_num - 1, res_num - 1):
        if(type(res[i][2]) is not str):
            print("          (" + str(res[i][0]) + ") " + str(res[i][1]) + "     Accuracy: %.2f%%" % (float(res[i][2]) * 100))
        else:
            print("          (" + str(res[i][0]) + ") " + str(res[i][1]) + "     Accuracy: " + res[i][2])


def save_and_exit():
    print("#########  EXITING: Writing results to \'selections.xlsx\'  #################")
    print(cat_strings)
    print(selections)

    global server

    df = pd.DataFrame({'input' : cat_strings,
                        'output' : selections})

    writer = pd.ExcelWriter("selections.xlsx", engine='openpyxl')
    df.to_excel(writer, "Sheet1")
    writer.save()
    del server
    exit()

def save():

    print("#########  AUTO SAVE: Writing results to \'selections.xlsx\'  #################")

    print(cat_strings)
    print(selections)

    df = pd.DataFrame({'input' : cat_strings,
                        'output' : selections})

    writer = pd.ExcelWriter("selections.xlsx", engine='openpyxl')
    df.to_excel(writer, "Sheet1")
    writer.save()


# DEAD CODE maybe used in the future?
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_wb():

    book = None
    fname = "selections.xlsx"
    if os.path.isfile(fname):
        book=load_workbook(fname)
    else:
        workbook2=xlwt.Workbook(fname)
        ws = workbook2.add_sheet('Sheet1')
        workbook2.save(fname)
        book = open_workbook(fname)

    return book

def update():
    server = Server()
    server.Update_PA_DB_w_Session_Data()
    del server


class Server:
    server = None
    connection = None
    mycursor = None
    pc = None

    def __init__(self, pc=None):
        self.pc = pc
        sshtunnel.SSH_TIMEOUT = 350.0
        sshtunnel.TUNNEL_TIMEOUT = 350.0
        self.server = sshtunnel.SSHTunnelForwarder(
            ('ssh.pythonanywhere.com'),
            ssh_username='iclam19', ssh_password='@astest@1234',
            remote_bind_address=('iclam19.mysql.pythonanywhere-services.com', 3306)
        )
        self.server.start()

    def connect(self):
        try:
            self.connection = mysql.connector.connect(
                user='iclam19', password='astest1234',
                host='127.0.0.1', port=self.server.local_bind_port,
                database='iclam19$AssembledSupply',
            )
            if(self.connection.is_connected()):
                self.mycursor = self.connection.cursor()
            else:
                print("Did not connect")
                return None

        except Error as e:
            print(e)


    def __del__(self):
        self.server.stop()


    def Update_PA_DB_w_Session_Data(self):
        df = pd.read_excel('selections.xlsx', sheet_name='Sheet1')
        self.connect()
        for index, row in df.iterrows():
            #print(row['input'], row['output'])
            i_c = row['input']
            o_c = row['output']
            i_c_end = i_c.rsplit('|',1)[-1]
            val = (str(o_c), str(i_c_end), str(o_c),str(i_c))
            sql = sql = "Update ft_ml_categories Set output_category = %s, output_category_UI = (Select * From (SELECT Distinct concat(replace(output_category_UI,substring_index(output_category_UI,'|',-1),''),%s) FROM iclam19$AssembledSupply.ft_ml_categories Where output_category = %s)q1) Where input_category = %s"
            self.mycursor.execute(sql, val)
            self.connection.commit()
            print(self.mycursor.rowcount, "# of record(s) updated")

        self.connection.close()


    def Select_Input_Cat_From_DB(self):
        #new session input_category data that only has the primary category the user selected
        self.connect()

        sql = "Select input_category From ft_ml_categories Where output_category is null and primary_category = %s Order by id asc limit 100"

        self.mycursor.execute(sql,(self.pc,))
        returned_items = list(self.mycursor.fetchall())
        self.connection.close()

        return returned_items


    def Select_Training_Data_From_DB(self):
         #new session input & output_category data that only has the primary category the user selected
        self.connect()

        sql = "Select input_category, output_category From ft_ml_categories Where output_category is not null and primary_category = %s"

        self.mycursor.execute(sql,(self.pc,))
        returned_items = list(self.mycursor.fetchall())
        self.connection.close()

        input_categories = []
        output_categories = []
        for cat in returned_items:
            input_categories.append(cat[0])
            output_categories.append(cat[1])

        return (input_categories, output_categories)


    def Select_Unique_Output_Cat_From_DB(self):

        self.connect()

        sql = "Select Distinct output_category FROM iclam19$AssembledSupply.ft_ml_categories Where output_category is not null and primary_category = %s"

        self.mycursor.execute(sql,(self.pc,))
        output_unique = list(self.mycursor.fetchall())
        self.connection.close()

        return [out[0] for out in output_unique]


if(__name__ == "__main__"):
    try:
        if(len(sys.argv) == 2):
            globals()[sys.argv[1]]()
        elif(len(sys.argv) == 3):
            globals()[sys.argv[1]](sys.argv[2])
        else:
            print("Usage: python categorizer.py method_name [args]")
            print("method_names:")
            print("     run_categorizer")
            print("     pickle_model")
    except KeyboardInterrupt:
        save_and_exit()
        try:
            sys.exit(0)
        except SystemExit:
            os._exit(0)
